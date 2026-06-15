import openpyxl, json

wb = openpyxl.load_workbook("C:/Users/nicol/Downloads/Anguilla Engineering Resources.xlsx")
ws = wb.active

tabs = []
current_tab = None
items = []

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    b = row[1]  # Column B
    if not b.value:
        continue
    
    is_bold = b.font.bold if b.font else False
    is_italic = b.font.italic if b.font else False
    
    if is_bold:
        if current_tab:
            tabs.append({'name': current_tab, 'items': items})
        current_tab = str(b.value)
        items = []
    else:
        items.append({
            'title': str(b.value),
            'best': is_italic,
            'link': b.hyperlink.target if b.hyperlink else '',
            'length': str(row[2].value or ''),
            'source': str(row[3].value or ''),
            'age':    str(row[4].value or ''),
            'ageNotes': str(row[5].value or ''),   # Column F
            'topic':    str(row[6].value or ''),   # Column G
            'use':      str(row[7].value or ''),   # Column H
            'teaches':  str(row[8].value or ''),   # Column I
            'notes':    str(row[9].value or ''),   # Column J
            'oces':     str(row[0].value or ''),
        })

tabs.append({'name': current_tab, 'items': items})
print(json.dumps(tabs))